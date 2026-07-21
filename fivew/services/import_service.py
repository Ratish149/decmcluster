import csv
import io
from decimal import Decimal

import openpyxl

from ..models import FiveWActivity


def to_int(val):
    if val is None:
        return 0
    if isinstance(val, int):
        return val
    if isinstance(val, float):
        return int(val)
    val_str = str(val).strip()
    if not val_str or val_str.lower() in ("none", "null", "na", "n/a", "-"):
        return 0
    try:
        return int(float(val_str))
    except ValueError:
        return 0


def to_decimal(val):
    if val is None:
        return None
    if isinstance(val, (int, float, Decimal)):
        return Decimal(str(val))
    val_str = str(val).strip()
    if not val_str or val_str.lower() in ("none", "null", "na", "n/a", "-"):
        return None
    try:
        return Decimal(val_str)
    except ValueError:
        return None


def to_bool(val):
    if val is None:
        return None
    if isinstance(val, bool):
        return val
    val_str = str(val).strip().lower()
    if val_str in ("yes", "y", "true", "1"):
        return True
    if val_str in ("no", "n", "false", "0"):
        return False
    return None


def to_str(val):
    if val is None:
        return None
    val_str = str(val).strip()
    if not val_str or val_str.lower() in ("none", "null", "na", "n/a", "-"):
        return None
    return val_str


def import_fivew_from_excel(file_path):
    """
    Parses Excel and directly creates FiveWActivity records.
    """
    if hasattr(file_path, "read"):
        wb = openpyxl.load_workbook(file_path, data_only=True)
    else:
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)

    if "5Ws" in wb.sheetnames:
        sheet = wb["5Ws"]
    else:
        sheet = wb.active

    created_count = 0

    # Headers at row 2, data starts at row 3
    for r in range(3, sheet.max_row + 1):
        # We check if there's any data in the row
        row_values = [sheet.cell(row=r, column=col).value for col in range(1, 61)]
        if not any(v is not None for v in row_values):
            continue

        reporting_org_name = sheet.cell(row=r, column=3).value
        # Check if reporting organization name is present, skip otherwise
        if not reporting_org_name or not str(reporting_org_name).strip():
            continue

        FiveWActivity.objects.create(
            donor=to_str(sheet.cell(row=r, column=1).value),
            donor_names=to_str(sheet.cell(row=r, column=2).value),
            reporting_org_name=to_str(reporting_org_name),
            ro_code=to_str(sheet.cell(row=r, column=4).value),
            reporting_org_type=to_str(sheet.cell(row=r, column=5).value),
            other_ip_name=to_str(sheet.cell(row=r, column=6).value),
            ip_code=to_str(sheet.cell(row=r, column=7).value),
            ip_type=to_str(sheet.cell(row=r, column=8).value),
            reporting_month=to_str(sheet.cell(row=r, column=9).value),
            activity_status=to_str(sheet.cell(row=r, column=10).value),
            state_abyei=to_str(sheet.cell(row=r, column=11).value),
            admin1_code=to_str(sheet.cell(row=r, column=12).value),
            province=to_str(sheet.cell(row=r, column=13).value),
            admin2_code=to_str(sheet.cell(row=r, column=14).value),
            admin_level_2_code=to_str(sheet.cell(row=r, column=15).value),
            admin3_code=to_str(sheet.cell(row=r, column=16).value),
            location_evac_name=to_str(sheet.cell(row=r, column=17).value),
            cluster_name=to_str(sheet.cell(row=r, column=18).value),
            hrp_non_hrp=to_str(sheet.cell(row=r, column=19).value),
            project_number=to_str(sheet.cell(row=r, column=20).value),
            project_name=to_str(sheet.cell(row=r, column=21).value),
            activity=to_str(sheet.cell(row=r, column=22).value),
            indicator=to_str(sheet.cell(row=r, column=23).value),
            unit=to_str(sheet.cell(row=r, column=24).value),
            target=to_int(sheet.cell(row=r, column=25).value),
            total_value=to_decimal(sheet.cell(row=r, column=26).value),
            new_beneficiaries=to_bool(sheet.cell(row=r, column=27).value),
            beneficiaries_type_under_18=to_str(sheet.cell(row=r, column=28).value),
            child_male_under_18=to_int(sheet.cell(row=r, column=29).value),
            child_female_under_18=to_int(sheet.cell(row=r, column=30).value),
            adult_male_18_60=to_int(sheet.cell(row=r, column=31).value),
            adult_female_18_60=to_int(sheet.cell(row=r, column=32).value),
            elderly_male_60_plus=to_int(sheet.cell(row=r, column=33).value),
            elderly_female_60_plus=to_int(sheet.cell(row=r, column=34).value),
            total_beneficiaries_reached=to_int(sheet.cell(row=r, column=35).value),
            people_with_disability=to_int(sheet.cell(row=r, column=36).value),
            is_mpc=to_bool(sheet.cell(row=r, column=37).value),
            modality=to_str(sheet.cell(row=r, column=38).value),
            type_of_modality=to_str(sheet.cell(row=r, column=39).value),
            delivery_mechanism=to_str(sheet.cell(row=r, column=40).value),
            number_of_transfers=to_int(sheet.cell(row=r, column=41).value),
            value_ssp=to_decimal(sheet.cell(row=r, column=42).value),
            value_usd=to_decimal(sheet.cell(row=r, column=43).value),
            comments=to_str(sheet.cell(row=r, column=44).value),
            contribute_hrp_aap=to_str(sheet.cell(row=r, column=45).value),
            hrp_aap_indicators=to_str(sheet.cell(row=r, column=46).value),
            activity_type=to_str(sheet.cell(row=r, column=47).value),
            sub_activity_type=to_str(sheet.cell(row=r, column=48).value),
            measurements=to_str(sheet.cell(row=r, column=49).value),
            achieved=to_int(sheet.cell(row=r, column=50).value),
            column1=to_str(sheet.cell(row=r, column=51).value),
            boys_above_5=to_int(sheet.cell(row=r, column=52).value),
            girls_above_5=to_int(sheet.cell(row=r, column=53).value),
            boys_5_17=to_int(sheet.cell(row=r, column=54).value),
            girls_5_17=to_int(sheet.cell(row=r, column=55).value),
            men_18_59=to_int(sheet.cell(row=r, column=56).value),
            women_18_59=to_int(sheet.cell(row=r, column=57).value),
            men_60_plus=to_int(sheet.cell(row=r, column=58).value),
            women_60_plus=to_int(sheet.cell(row=r, column=59).value),
            total_reached_quarter=to_int(sheet.cell(row=r, column=60).value),
        )
        created_count += 1

    return created_count, 0


def import_fivew_from_csv(file_obj):
    """
    Parses CSV and directly creates FiveWActivity records.
    """
    content = file_obj.read()
    if isinstance(content, bytes):
        try:
            decoded_content = content.decode("utf-8")
        except UnicodeDecodeError:
            decoded_content = content.decode("latin-1")
    else:
        decoded_content = content

    csv_file = io.StringIO(decoded_content)
    reader = csv.reader(csv_file)
    rows = list(reader)

    # Find the header row (typically row starts with Donor or Donors Names)
    header_row_index = -1
    for idx, row in enumerate(rows[:5]):
        if len(row) > 2 and "donor" in row[0].strip().lower():
            header_row_index = idx
            break

    if header_row_index == -1:
        # Default fallback
        header_row_index = 1

    data_rows = rows[header_row_index + 1 :]

    created_count = 0

    for row in data_rows:
        if not row or len(row) < 3:
            continue

        # Pad row with None to avoid index errors
        padded_row = row + [None] * (65 - len(row))

        reporting_org_name = padded_row[2]
        if not reporting_org_name or not str(reporting_org_name).strip():
            continue

        FiveWActivity.objects.create(
            donor=to_str(padded_row[0]),
            donor_names=to_str(padded_row[1]),
            reporting_org_name=to_str(reporting_org_name),
            ro_code=to_str(padded_row[3]),
            reporting_org_type=to_str(padded_row[4]),
            other_ip_name=to_str(padded_row[5]),
            ip_code=to_str(padded_row[6]),
            ip_type=to_str(padded_row[7]),
            reporting_month=to_str(padded_row[8]),
            activity_status=to_str(padded_row[9]),
            state_abyei=to_str(padded_row[10]),
            admin1_code=to_str(padded_row[11]),
            province=to_str(padded_row[12]),
            admin2_code=to_str(padded_row[13]),
            admin_level_2_code=to_str(padded_row[14]),
            admin3_code=to_str(padded_row[15]),
            location_evac_name=to_str(padded_row[16]),
            cluster_name=to_str(padded_row[17]),
            hrp_non_hrp=to_str(padded_row[18]),
            project_number=to_str(padded_row[19]),
            project_name=to_str(padded_row[20]),
            activity=to_str(padded_row[21]),
            indicator=to_str(padded_row[22]),
            unit=to_str(padded_row[23]),
            target=to_int(padded_row[24]),
            total_value=to_decimal(padded_row[25]),
            new_beneficiaries=to_bool(padded_row[26]),
            beneficiaries_type_under_18=to_str(padded_row[27]),
            child_male_under_18=to_int(padded_row[28]),
            child_female_under_18=to_int(padded_row[29]),
            adult_male_18_60=to_int(padded_row[30]),
            adult_female_18_60=to_int(padded_row[31]),
            elderly_male_60_plus=to_int(padded_row[32]),
            elderly_female_60_plus=to_int(padded_row[33]),
            total_beneficiaries_reached=to_int(padded_row[34]),
            people_with_disability=to_int(padded_row[35]),
            is_mpc=to_bool(padded_row[36]),
            modality=to_str(padded_row[37]),
            type_of_modality=to_str(padded_row[38]),
            delivery_mechanism=to_str(padded_row[39]),
            number_of_transfers=to_int(padded_row[40]),
            value_ssp=to_decimal(padded_row[41]),
            value_usd=to_decimal(padded_row[42]),
            comments=to_str(padded_row[43]),
            contribute_hrp_aap=to_str(padded_row[44]),
            hrp_aap_indicators=to_str(padded_row[45]),
            activity_type=to_str(padded_row[46]),
            sub_activity_type=to_str(padded_row[47]),
            measurements=to_str(padded_row[48]),
            achieved=to_int(padded_row[49]),
            column1=to_str(padded_row[50]),
            boys_above_5=to_int(padded_row[51]),
            girls_above_5=to_int(padded_row[52]),
            boys_5_17=to_int(padded_row[53]),
            girls_5_17=to_int(padded_row[54]),
            men_18_59=to_int(padded_row[55]),
            women_18_59=to_int(padded_row[56]),
            men_60_plus=to_int(padded_row[57]),
            women_60_plus=to_int(padded_row[58]),
            total_reached_quarter=to_int(padded_row[59]),
        )
        created_count += 1

    return created_count, 0
