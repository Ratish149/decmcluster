import csv
import datetime
import io
from decimal import Decimal

import openpyxl

from ..models import VillageAssessment


def to_int(val):
    if val is None:
        return 0
    if isinstance(val, int):
        return val
    if isinstance(val, float):
        return int(val)
    val_str = str(val).strip()
    if not val_str or val_str.lower() in ("none", "null", "na", "n/a", "-"):
        return 0
    try:
        return int(float(val_str))
    except ValueError:
        return 0


def to_decimal(val):
    if val is None:
        return None
    if isinstance(val, (int, float, Decimal)):
        return Decimal(str(val))
    val_str = str(val).strip()
    if not val_str or val_str.lower() in ("none", "null", "na", "n/a", "-"):
        return None
    try:
        return Decimal(val_str)
    except ValueError:
        return None


def to_bool(val):
    if val is None:
        return None
    if isinstance(val, bool):
        return val
    val_str = str(val).strip().lower()
    if val_str in ("yes", "y", "true", "1"):
        return True
    if val_str in ("no", "n", "false", "0"):
        return False
    return None


def to_str(val):
    if val is None:
        return None
    val_str = str(val).strip()
    if not val_str or val_str.lower() in ("none", "null", "na", "n/a", "-"):
        return None
    return val_str


def parse_date(val):
    if val is None:
        return None
    if isinstance(val, (datetime.date, datetime.datetime)):
        return val.date() if isinstance(val, datetime.datetime) else val
    if isinstance(val, (int, float)):
        try:
            return datetime.date(1899, 12, 30) + datetime.timedelta(days=int(val))
        except Exception:
            return None
    val_str = str(val).strip()
    if not val_str or val_str.lower() in ("none", "null", "na", "n/a", "-"):
        return None
    for fmt in (
        "%Y-%m-%dT%H:%M:%S",
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d",
        "%d/%m/%Y",
        "%m/%d/%Y",
    ):
        try:
            return datetime.datetime.strptime(val_str, fmt).date()
        except ValueError:
            continue
    return None


def parse_time(val):
    if val is None:
        return None
    if isinstance(val, datetime.time):
        return val
    if isinstance(val, datetime.datetime):
        return val.time()
    val_str = str(val).strip()
    if not val_str or val_str.lower() in ("none", "null", "na", "n/a", "-"):
        return None
    for fmt in ("%H:%M:%S", "%H:%M"):
        try:
            return datetime.datetime.strptime(val_str, fmt).time()
        except ValueError:
            continue
    return None


def parse_datetime(val):
    if val is None:
        return None
    if isinstance(val, datetime.datetime):
        return val
    if isinstance(val, datetime.date):
        return datetime.datetime.combine(val, datetime.time.min)
    if isinstance(val, (int, float)):
        try:
            return datetime.datetime(1899, 12, 30) + datetime.timedelta(days=float(val))
        except Exception:
            return None
    val_str = str(val).strip()
    if not val_str or val_str.lower() in ("none", "null", "na", "n/a", "-"):
        return None
    for fmt in (
        "%Y-%m-%dT%H:%M:%S",
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d %H:%M",
        "%d/%m/%Y %H:%M",
        "%m/%d/%Y %H:%M",
    ):
        try:
            return datetime.datetime.strptime(val_str, fmt)
        except ValueError:
            continue
    return None


def import_village_assessments_from_excel(file_path):
    """
    Parses Excel and directly creates VillageAssessment records.
    """
    if hasattr(file_path, "read"):
        wb = openpyxl.load_workbook(file_path, data_only=True)
    else:
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    sheet = wb.active

    created_count = 0

    # Headers at row 1, data starts at row 2
    for r in range(2, sheet.max_row + 1):
        # We check if there's any data in the row
        row_values = [sheet.cell(row=r, column=col).value for col in range(1, 114)]
        if not any(v is not None for v in row_values):
            continue

        village_name = sheet.cell(row=r, column=54).value
        # Check if village name is present
        if not village_name or not str(village_name).strip():
            continue

        VillageAssessment.objects.create(
            survey_start=parse_date(sheet.cell(row=r, column=1).value),
            survey_end=parse_date(sheet.cell(row=r, column=2).value),
            survey_date=parse_date(sheet.cell(row=r, column=3).value),
            enumerator_username=to_str(sheet.cell(row=r, column=4).value),
            device_id=to_str(sheet.cell(row=r, column=5).value),
            audit_file=to_str(sheet.cell(row=r, column=6).value),
            audit_url=to_str(sheet.cell(row=r, column=7).value),
            consent=to_str(sheet.cell(row=r, column=8).value),
            methodology_individual_ki=to_str(sheet.cell(row=r, column=9).value),
            methodology_group_ki=to_str(sheet.cell(row=r, column=10).value),
            methodology_direct_observation=to_str(sheet.cell(row=r, column=11).value),
            methodology_other=to_str(sheet.cell(row=r, column=12).value),
            data_collection_method=to_str(sheet.cell(row=r, column=13).value),
            ki1_name=to_str(sheet.cell(row=r, column=14).value),
            ki1_type=to_str(sheet.cell(row=r, column=15).value),
            ki1_gender=to_str(sheet.cell(row=r, column=16).value),
            ki1_age=to_int(sheet.cell(row=r, column=17).value),
            ki1_contact=to_str(sheet.cell(row=r, column=18).value),
            ki2_name=to_str(sheet.cell(row=r, column=19).value),
            ki2_type=to_str(sheet.cell(row=r, column=20).value),
            ki2_gender=to_str(sheet.cell(row=r, column=21).value),
            ki2_age=to_int(sheet.cell(row=r, column=22).value),
            ki2_contact=to_str(sheet.cell(row=r, column=23).value),
            ki3_name=to_str(sheet.cell(row=r, column=24).value),
            ki3_type=to_str(sheet.cell(row=r, column=25).value),
            ki3_gender=to_str(sheet.cell(row=r, column=26).value),
            ki3_age=to_int(sheet.cell(row=r, column=27).value),
            ki3_contact=to_str(sheet.cell(row=r, column=28).value),
            ki4_name=to_str(sheet.cell(row=r, column=29).value),
            ki4_type=to_str(sheet.cell(row=r, column=30).value),
            ki4_gender=to_str(sheet.cell(row=r, column=31).value),
            ki4_age=to_int(sheet.cell(row=r, column=32).value),
            ki4_contact=to_str(sheet.cell(row=r, column=33).value),
            ki5_name=to_str(sheet.cell(row=r, column=34).value),
            ki5_type=to_str(sheet.cell(row=r, column=35).value),
            ki5_gender=to_str(sheet.cell(row=r, column=36).value),
            ki5_age=to_int(sheet.cell(row=r, column=37).value),
            ki5_contact=to_str(sheet.cell(row=r, column=38).value),
            ki6_name=to_str(sheet.cell(row=r, column=39).value),
            ki6_type=to_str(sheet.cell(row=r, column=40).value),
            ki6_gender=to_str(sheet.cell(row=r, column=41).value),
            ki6_age=to_int(sheet.cell(row=r, column=42).value),
            ki6_contact=to_str(sheet.cell(row=r, column=43).value),
            assessment_date=parse_date(sheet.cell(row=r, column=44).value),
            assessment_start_time=parse_time(sheet.cell(row=r, column=45).value),
            enumerator1_name=to_str(sheet.cell(row=r, column=46).value),
            enumerator1_phone=to_str(sheet.cell(row=r, column=47).value),
            enumerator1_gender=to_str(sheet.cell(row=r, column=48).value),
            enumerator2_name=to_str(sheet.cell(row=r, column=49).value),
            enumerator2_phone=to_str(sheet.cell(row=r, column=50).value),
            enumerator2_gender=to_str(sheet.cell(row=r, column=51).value),
            province=to_str(sheet.cell(row=r, column=52).value),
            area_council=to_str(sheet.cell(row=r, column=53).value),
            village_name=to_str(village_name),
            village_other=to_str(sheet.cell(row=r, column=55).value),
            village_condition=to_str(sheet.cell(row=r, column=56).value),
            idp_present=to_bool(sheet.cell(row=r, column=57).value),
            idp_households_total=to_int(sheet.cell(row=r, column=58).value),
            idp_infant_male=to_int(sheet.cell(row=r, column=59).value),
            idp_infant_female=to_int(sheet.cell(row=r, column=60).value),
            idp_child_1_5_male=to_int(sheet.cell(row=r, column=61).value),
            idp_child_1_5_female=to_int(sheet.cell(row=r, column=62).value),
            idp_child_6_12_male=to_int(sheet.cell(row=r, column=63).value),
            idp_child_6_12_female=to_int(sheet.cell(row=r, column=64).value),
            idp_adolescent_male=to_int(sheet.cell(row=r, column=65).value),
            idp_adolescent_female=to_int(sheet.cell(row=r, column=66).value),
            idp_adult_male=to_int(sheet.cell(row=r, column=67).value),
            idp_adult_female=to_int(sheet.cell(row=r, column=68).value),
            idp_elderly_male=to_int(sheet.cell(row=r, column=69).value),
            idp_elderly_female=to_int(sheet.cell(row=r, column=70).value),
            idp_male_total=to_int(sheet.cell(row=r, column=71).value),
            idp_female_total=to_int(sheet.cell(row=r, column=72).value),
            idp_individuals_total=to_int(sheet.cell(row=r, column=73).value),
            returnees_present=to_bool(sheet.cell(row=r, column=74).value),
            returnee_households_total=to_int(sheet.cell(row=r, column=75).value),
            returnee_individuals_total=to_int(sheet.cell(row=r, column=76).value),
            pregnant_women_count=to_int(sheet.cell(row=r, column=77).value),
            female_headed_hh=to_int(sheet.cell(row=r, column=78).value),
            elderly_headed_hh=to_int(sheet.cell(row=r, column=79).value),
            male_headed_hh=to_int(sheet.cell(row=r, column=80).value),
            child_headed_hh=to_int(sheet.cell(row=r, column=81).value),
            pwd_total=to_int(sheet.cell(row=r, column=82).value),
            idp_pwd_total=to_int(sheet.cell(row=r, column=83).value),
            shelter_primary=to_str(sheet.cell(row=r, column=84).value),
            shelter_secondary=to_str(sheet.cell(row=r, column=85).value),
            displacement_shelter_type=to_str(sheet.cell(row=r, column=86).value),
            displaced_hh_estimated=to_int(sheet.cell(row=r, column=87).value),
            displacement_duration=to_str(sheet.cell(row=r, column=88).value),
            housing_type_pre_cyclone=to_str(sheet.cell(row=r, column=89).value),
            house_rebuild_duration=to_str(sheet.cell(row=r, column=90).value),
            rebuild_material_type=to_str(sheet.cell(row=r, column=91).value),
            house_cyclone_resilience=to_str(sheet.cell(row=r, column=92).value),
            remaining_idp_intention=to_str(sheet.cell(row=r, column=93).value),
            seasonal_worker_level=to_str(sheet.cell(row=r, column=94).value),
            community_participation=to_str(sheet.cell(row=r, column=95).value),
            cdccc_exists=to_bool(sheet.cell(row=r, column=96).value),
            early_warning_received=to_bool(sheet.cell(row=r, column=97).value),
            annual_population_displaced=to_int(sheet.cell(row=r, column=98).value),
            top_need_1=to_str(sheet.cell(row=r, column=99).value),
            top_need_2=to_str(sheet.cell(row=r, column=100).value),
            top_need_3=to_str(sheet.cell(row=r, column=101).value),
            gps_latitude=to_decimal(sheet.cell(row=r, column=102).value),
            gps_longitude=to_decimal(sheet.cell(row=r, column=103).value),
            gps_altitude=to_decimal(sheet.cell(row=r, column=104).value),
            gps_precision=to_decimal(sheet.cell(row=r, column=105).value),
            record_id=to_str(sheet.cell(row=r, column=106).value),
            record_uuid=to_str(sheet.cell(row=r, column=107).value),
            submission_time=parse_datetime(sheet.cell(row=r, column=108).value),
            validation_status=to_str(sheet.cell(row=r, column=109).value),
            submission_status=to_str(sheet.cell(row=r, column=110).value),
            submitted_by=to_str(sheet.cell(row=r, column=111).value),
            form_version=to_str(sheet.cell(row=r, column=112).value),
            record_index=to_int(sheet.cell(row=r, column=113).value),
        )
        created_count += 1

    return created_count, 0


def import_village_assessments_from_csv(file_obj):
    """
    Parses CSV and directly creates VillageAssessment records.
    """
    content = file_obj.read()
    if isinstance(content, bytes):
        try:
            decoded_content = content.decode("utf-8")
        except UnicodeDecodeError:
            decoded_content = content.decode("latin-1")
    else:
        decoded_content = content

    csv_file = io.StringIO(decoded_content)
    reader = csv.reader(csv_file)
    rows = list(reader)

    # Find the header row
    header_row_index = -1
    for idx, row in enumerate(rows[:5]):
        if len(row) > 2 and "survey_start" in row[0].strip().lower():
            header_row_index = idx
            break

    if header_row_index == -1:
        header_row_index = 0

    data_rows = rows[header_row_index + 1 :]

    created_count = 0

    for row in data_rows:
        if not row or len(row) < 54:
            continue

        padded_row = row + [None] * (115 - len(row))

        village_name = padded_row[53]
        if not village_name or not str(village_name).strip():
            continue

        VillageAssessment.objects.create(
            survey_start=parse_date(padded_row[0]),
            survey_end=parse_date(padded_row[1]),
            survey_date=parse_date(padded_row[2]),
            enumerator_username=to_str(padded_row[3]),
            device_id=to_str(padded_row[4]),
            audit_file=to_str(padded_row[5]),
            audit_url=to_str(padded_row[6]),
            consent=to_str(padded_row[7]),
            methodology_individual_ki=to_str(padded_row[8]),
            methodology_group_ki=to_str(padded_row[9]),
            methodology_direct_observation=to_str(padded_row[10]),
            methodology_other=to_str(padded_row[11]),
            data_collection_method=to_str(padded_row[12]),
            ki1_name=to_str(padded_row[13]),
            ki1_type=to_str(padded_row[14]),
            ki1_gender=to_str(padded_row[15]),
            ki1_age=to_int(padded_row[16]),
            ki1_contact=to_str(padded_row[17]),
            ki2_name=to_str(padded_row[18]),
            ki2_type=to_str(padded_row[19]),
            ki2_gender=to_str(padded_row[20]),
            ki2_age=to_int(padded_row[21]),
            ki2_contact=to_str(padded_row[22]),
            ki3_name=to_str(padded_row[23]),
            ki3_type=to_str(padded_row[24]),
            ki3_gender=to_str(padded_row[25]),
            ki3_age=to_int(padded_row[26]),
            ki3_contact=to_str(padded_row[27]),
            ki4_name=to_str(padded_row[28]),
            ki4_type=to_str(padded_row[29]),
            ki4_gender=to_str(padded_row[30]),
            ki4_age=to_int(padded_row[31]),
            ki4_contact=to_str(padded_row[32]),
            ki5_name=to_str(padded_row[33]),
            ki5_type=to_str(padded_row[34]),
            ki5_gender=to_str(padded_row[35]),
            ki5_age=to_int(padded_row[36]),
            ki5_contact=to_str(padded_row[37]),
            ki6_name=to_str(padded_row[38]),
            ki6_type=to_str(padded_row[39]),
            ki6_gender=to_str(padded_row[40]),
            ki6_age=to_int(padded_row[41]),
            ki6_contact=to_str(padded_row[42]),
            assessment_date=parse_date(padded_row[43]),
            assessment_start_time=parse_time(padded_row[44]),
            enumerator1_name=to_str(padded_row[45]),
            enumerator1_phone=to_str(padded_row[46]),
            enumerator1_gender=to_str(padded_row[47]),
            enumerator2_name=to_str(padded_row[48]),
            enumerator2_phone=to_str(padded_row[49]),
            enumerator2_gender=to_str(padded_row[50]),
            province=to_str(padded_row[51]),
            area_council=to_str(padded_row[52]),
            village_name=to_str(village_name),
            village_other=to_str(padded_row[54]),
            village_condition=to_str(padded_row[55]),
            idp_present=to_bool(padded_row[56]),
            idp_households_total=to_int(padded_row[57]),
            idp_infant_male=to_int(padded_row[58]),
            idp_infant_female=to_int(padded_row[59]),
            idp_child_1_5_male=to_int(padded_row[60]),
            idp_child_1_5_female=to_int(padded_row[61]),
            idp_child_6_12_male=to_int(padded_row[62]),
            idp_child_6_12_female=to_int(padded_row[63]),
            idp_adolescent_male=to_int(padded_row[64]),
            idp_adolescent_female=to_int(padded_row[65]),
            idp_adult_male=to_int(padded_row[66]),
            idp_adult_female=to_int(padded_row[67]),
            idp_elderly_male=to_int(padded_row[68]),
            idp_elderly_female=to_int(padded_row[69]),
            idp_male_total=to_int(padded_row[70]),
            idp_female_total=to_int(padded_row[71]),
            idp_individuals_total=to_int(padded_row[72]),
            returnees_present=to_bool(padded_row[73]),
            returnee_households_total=to_int(padded_row[74]),
            returnee_individuals_total=to_int(padded_row[75]),
            pregnant_women_count=to_int(padded_row[76]),
            female_headed_hh=to_int(padded_row[77]),
            elderly_headed_hh=to_int(padded_row[78]),
            male_headed_hh=to_int(padded_row[79]),
            child_headed_hh=to_int(padded_row[80]),
            pwd_total=to_int(padded_row[81]),
            idp_pwd_total=to_int(padded_row[82]),
            shelter_primary=to_str(padded_row[83]),
            shelter_secondary=to_str(padded_row[84]),
            displacement_shelter_type=to_str(padded_row[85]),
            displaced_hh_estimated=to_int(padded_row[86]),
            displacement_duration=to_str(padded_row[87]),
            housing_type_pre_cyclone=to_str(padded_row[88]),
            house_rebuild_duration=to_str(padded_row[89]),
            rebuild_material_type=to_str(padded_row[90]),
            house_cyclone_resilience=to_str(padded_row[91]),
            remaining_idp_intention=to_str(padded_row[92]),
            seasonal_worker_level=to_str(padded_row[93]),
            community_participation=to_str(padded_row[94]),
            cdccc_exists=to_bool(padded_row[95]),
            early_warning_received=to_bool(padded_row[96]),
            annual_population_displaced=to_int(padded_row[97]),
            top_need_1=to_str(padded_row[98]),
            top_need_2=to_str(padded_row[99]),
            top_need_3=to_str(padded_row[100]),
            gps_latitude=to_decimal(padded_row[101]),
            gps_longitude=to_decimal(padded_row[102]),
            gps_altitude=to_decimal(padded_row[103]),
            gps_precision=to_decimal(padded_row[104]),
            record_id=to_str(padded_row[105]),
            record_uuid=to_str(padded_row[106]),
            submission_time=parse_datetime(padded_row[107]),
            validation_status=to_str(padded_row[108]),
            submission_status=to_str(padded_row[109]),
            submitted_by=to_str(padded_row[110]),
            form_version=to_str(padded_row[111]),
            record_index=to_int(padded_row[112]),
        )
        created_count += 1

    return created_count, 0
